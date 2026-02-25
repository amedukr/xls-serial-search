import os
import re
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


SUPPORTED_EXT = (".xls",)


def clean_text(s: str) -> str:
    """
    Нормалізація для пошуку:
    - прибирає лапки різних типів
    - прибирає всі пробіли/таби/переноси (щоб серійник з пробілами знаходився)
    - lower()
    """
    if s is None:
        return ""
    s = str(s)

    s = s.replace("\ufeff", "")       # BOM
    s = s.replace("\u00a0", " ")      # NBSP
    s = s.replace("\t", " ")

    s = s.replace('"', "").replace("'", "")
    s = s.replace("“", "").replace("”", "").replace("«", "").replace("»", "")

    s = re.sub(r"\s+", "", s)
    return s.strip().lower()


def display_clean(s: str) -> str:
    """Як показувати серійник у результатах: без лапок/пробілів, але без lower()."""
    if s is None:
        return ""
    s = str(s)

    s = s.replace("\ufeff", "")
    s = s.replace("\u00a0", " ")
    s = s.replace("\t", " ")

    s = s.replace('"', "").replace("'", "")
    s = s.replace("“", "").replace("”", "").replace("«", "").replace("»", "")

    s = re.sub(r"\s+", "", s)
    return s.strip()


def iter_xls_files(root: str, recursive: bool = True):
    if recursive:
        for dirpath, _, filenames in os.walk(root):
            for name in filenames:
                if name.lower().endswith(SUPPORTED_EXT) and not name.startswith("~$"):
                    yield os.path.join(dirpath, name)
    else:
        for name in os.listdir(root):
            path = os.path.join(root, name)
            if os.path.isfile(path) and name.lower().endswith(SUPPORTED_EXT) and not name.startswith("~$"):
                yield path


def cell_to_string(book, sheet, r, c) -> str:
    cell = sheet.cell(r, c)
    v = cell.value
    if v is None:
        return ""

    if cell.ctype == xlrd.XL_CELL_NUMBER:
        try:
            fv = float(v)
            if fv.is_integer():
                return str(int(fv))
            return str(fv)
        except Exception:
            return str(v)

    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            y, m, d, hh, mm, ss = xlrd.xldate_as_tuple(v, book.datemode)
            if hh == mm == ss == 0:
                return f"{d:02d}.{m:02d}.{y:04d}"
            return f"{d:02d}.{m:02d}.{y:04d} {hh:02d}:{mm:02d}:{ss:02d}"
        except Exception:
            return str(v)

    return str(v)


def load_queries(single_query: str, query_file: str):
    """
    Повертає:
    - order_norm: список нормалізованих запитів (у порядку)
    - display_map: norm -> відображення в таблиці
    """
    display_map = {}
    order_norm = []

    def add_one(q_raw: str):
        disp = display_clean(q_raw)
        norm = clean_text(q_raw)
        if not norm:
            return
        if norm not in display_map:
            display_map[norm] = disp if disp else norm
            order_norm.append(norm)

    if single_query and single_query.strip():
        add_one(single_query.strip())

    if query_file and os.path.isfile(query_file):
        with open(query_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    add_one(line)

    return order_norm, display_map


def find_first_matches(folder: str, queries_norm: list[str], recursive: bool,
                       progress_cb=None, stop_flag=None):
    """
    Шукає перший файл для кожного серійника.
    Повертає found_map: q_norm -> file_path
    """
    found = {}
    remaining = set(queries_norm)

    files = list(iter_xls_files(folder, recursive=recursive))
    total = len(files)

    for idx, fpath in enumerate(files, 1):
        if stop_flag and stop_flag.is_set():
            break

        if progress_cb:
            progress_cb(idx, total, os.path.basename(fpath), len(found), len(queries_norm))

        if not remaining:
            break

        try:
            book = xlrd.open_workbook(fpath, on_demand=True)
        except Exception:
            continue

        try:
            for si in range(book.nsheets):
                if not remaining:
                    break
                sh = book.sheet_by_index(si)

                for r in range(sh.nrows):
                    if not remaining:
                        break
                    for c in range(sh.ncols):
                        raw = cell_to_string(book, sh, r, c)
                        if not raw:
                            continue
                        norm_cell = clean_text(raw)
                        if not norm_cell:
                            continue

                        hit_now = []
                        for q in remaining:
                            if q in norm_cell:
                                found[q] = fpath
                                hit_now.append(q)
                        for q in hit_now:
                            remaining.discard(q)

                        if not remaining:
                            break
        finally:
            try:
                book.release_resources()
            except Exception:
                pass

    return found


def save_results_xlsx(out_path: str, queries_order: list[str], display_map: dict, found_map: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    ws["A1"] = "Серійник"
    ws["B1"] = "Назва файлу"
    ws["C1"] = "Шлях (посилання)"
    for cell in ("A1", "B1", "C1"):
        ws[cell].font = Font(bold=True)

    link_font = Font(color="0000FF", underline="single")

    row = 2
    for q in queries_order:
        serial_disp = display_map.get(q, q)
        if q in found_map:
            fpath = found_map[q]
            fname = os.path.basename(fpath)

            ws.cell(row=row, column=1, value=serial_disp)
            ws.cell(row=row, column=2, value=fname)

            c = ws.cell(row=row, column=3, value=fpath)
            c.hyperlink = fpath
            c.font = link_font
        else:
            ws.cell(row=row, column=1, value=serial_disp)
            ws.cell(row=row, column=2, value="НЕ ЗНАЙДЕНО")
            ws.cell(row=row, column=3, value="")
        row += 1

    for col in range(1, 4):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(14, max_len + 2), 90)

    wb.save(out_path)


def open_file_crossplatform(path: str):
    # Windows/macOS/Linux
    try:
        if os.name == "nt":
            os.startfile(path)  # noqa
        else:
            import subprocess
            if sys.platform == "darwin":
                subprocess.run(["open", path], check=False)
            else:
                subprocess.run(["xdg-open", path], check=False)
    except Exception:
        pass


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Пошук серійників у .XLS")
        self.geometry("980x620")

        self.q = queue.Queue()
        self.stop_flag = threading.Event()
        self.worker = None
        self.last_results = None  # (queries_order, display_map, found_map)

        frm = ttk.Frame(self, padding=10)
        frm.pack(fill="x")

        self.var_folder = tk.StringVar()
        self.var_listfile = tk.StringVar()
        self.var_single = tk.StringVar()
        self.var_recursive = tk.BooleanVar(value=True)

        ttk.Label(frm, text="Папка з .xls:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.var_folder, width=70).grid(row=0, column=1, sticky="we", padx=6)
        ttk.Button(frm, text="Обрати...", command=self.pick_folder).grid(row=0, column=2, padx=4)

        ttk.Label(frm, text="Файл серійників (.txt):").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(frm, textvariable=self.var_listfile, width=70).grid(row=1, column=1, sticky="we", padx=6, pady=(6, 0))
        ttk.Button(frm, text="Обрати...", command=self.pick_listfile).grid(row=1, column=2, padx=4, pady=(6, 0))

        ttk.Label(frm, text="Або один серійник:").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(frm, textvariable=self.var_single, width=70).grid(row=2, column=1, sticky="we", padx=6, pady=(6, 0))

        ttk.Checkbutton(frm, text="Шукати у підпапках", variable=self.var_recursive).grid(row=2, column=2, sticky="w", pady=(6, 0))

        frm.columnconfigure(1, weight=1)

        btns = ttk.Frame(self, padding=(10, 0, 10, 10))
        btns.pack(fill="x")

        self.btn_start = ttk.Button(btns, text="Пошук", command=self.start_search)
        self.btn_stop = ttk.Button(btns, text="Стоп", command=self.stop_search, state="disabled")
        self.btn_save = ttk.Button(btns, text="Зберегти results.xlsx", command=self.save_xlsx, state="disabled")

        self.btn_start.pack(side="left")
        self.btn_stop.pack(side="left", padx=8)
        self.btn_save.pack(side="left")

        prog = ttk.Frame(self, padding=(10, 0, 10, 10))
        prog.pack(fill="x")

        self.pb = ttk.Progressbar(prog, mode="determinate")
        self.pb.pack(fill="x")
        self.lbl_status = ttk.Label(prog, text="Готово.")
        self.lbl_status.pack(anchor="w", pady=(6, 0))

        table_frame = ttk.Frame(self, padding=10)
        table_frame.pack(fill="both", expand=True)

        cols = ("serial", "filename", "path")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings")
        self.tree.heading("serial", text="Серійник")
        self.tree.heading("filename", text="Назва файлу")
        self.tree.heading("path", text="Шлях (подвійний клік — відкрити)")

        self.tree.column("serial", width=220, anchor="w")
        self.tree.column("filename", width=250, anchor="w")
        self.tree.column("path", width=460, anchor="w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.tree.bind("<Double-1>", self.on_double_click)

        self.after(100, self.poll_queue)

    def pick_folder(self):
        path = filedialog.askdirectory()
        if path:
            self.var_folder.set(path)

    def pick_listfile(self):
        path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if path:
            self.var_listfile.set(path)

    def set_status(self, text):
        self.lbl_status.config(text=text)

    def start_search(self):
        folder = self.var_folder.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("Увага", "Обери правильну папку з .xls файлами.")
            return

        qfile = self.var_listfile.get().strip()
        single = self.var_single.get().strip()

        queries_order, display_map = load_queries(single, qfile)
        if not queries_order:
            messagebox.showwarning("Увага", "Задай серійники: або файл .txt, або один серійник у полі.")
            return

        # очистити таблицю
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.last_results = None
        self.btn_save.config(state="disabled")

        self.stop_flag.clear()
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.pb["value"] = 0
        self.pb["maximum"] = 100
        self.set_status("Старт...")

        recursive = bool(self.var_recursive.get())

        def progress_cb(i, total, fname, found_cnt, total_q):
            self.q.put(("progress", i, total, fname, found_cnt, total_q))

        def worker():
            try:
                found_map = find_first_matches(
                    folder=folder,
                    queries_norm=queries_order,
                    recursive=recursive,
                    progress_cb=progress_cb,
                    stop_flag=self.stop_flag
                )
                self.q.put(("done", queries_order, display_map, found_map))
            except Exception as e:
                self.q.put(("error", str(e)))

        self.worker = threading.Thread(target=worker, daemon=True)
        self.worker.start()

    def stop_search(self):
        self.stop_flag.set()
        self.set_status("Зупиняю...")
        self.btn_stop.config(state="disabled")

    def save_xlsx(self):
        if not self.last_results:
            return
        queries_order, display_map, found_map = self.last_results

        out_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")],
            initialfile="results.xlsx"
        )
        if not out_path:
            return

        try:
            save_results_xlsx(out_path, queries_order, display_map, found_map)
            messagebox.showinfo("Готово", f"Збережено:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Помилка", str(e))

    def on_double_click(self, event):
        item = self.tree.selection()
        if not item:
            return
        values = self.tree.item(item[0], "values")
        if len(values) >= 3:
            path = values[2]
            if path and os.path.isfile(path):
                try:
                    os.startfile(path)  # Windows
                except Exception:
                    messagebox.showerror("Помилка", f"Не вдалося відкрити:\n{path}")

    def poll_queue(self):
        try:
            while True:
                msg = self.q.get_nowait()

                if msg[0] == "progress":
                    i, total, fname, found_cnt, total_q = msg[1:]
                    pct = int((i / max(total, 1)) * 100)
                    self.pb["value"] = pct
                    self.set_status(f"Файл {i}/{total}: {fname} | знайдено {found_cnt}/{total_q}")

                elif msg[0] == "done":
                    queries_order, display_map, found_map = msg[1:]
                    self.last_results = (queries_order, display_map, found_map)

                    for q in queries_order:
                        serial_disp = display_map.get(q, q)
                        if q in found_map:
                            fpath = found_map[q]
                            fname = os.path.basename(fpath)
                            self.tree.insert("", "end", values=(serial_disp, fname, fpath))
                        else:
                            self.tree.insert("", "end", values=(serial_disp, "НЕ ЗНАЙДЕНО", ""))

                    self.pb["value"] = 100
                    found_cnt = len(found_map)
                    total_q = len(queries_order)
                    self.set_status(f"Готово. Знайдено {found_cnt} з {total_q}.")
                    self.btn_save.config(state="normal")
                    self.btn_start.config(state="normal")
                    self.btn_stop.config(state="disabled")

                elif msg[0] == "error":
                    err = msg[1]
                    messagebox.showerror("Помилка", err)
                    self.btn_start.config(state="normal")
                    self.btn_stop.config(state="disabled")
                    self.set_status("Помилка.")
        except queue.Empty:
            pass

        self.after(100, self.poll_queue)


if __name__ == "__main__":
    App().mainloop()